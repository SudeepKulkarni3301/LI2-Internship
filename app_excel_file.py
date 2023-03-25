from flask import Flask, render_template, jsonify
import serial
import sys
import threading
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import datetime

app = Flask(__name__)

latest_tag = ''
tag_count = 0
tag_data = []

def convert_tag_from_bytes_to_hex(tag_bytes_list):
    tag_hex_value = ""
    for index, bytes_value in enumerate(tag_bytes_list):
        #   First 3 bytes and last byte are placeholders
        if index > 3 and index < 16:
            tag_hex_value += "{0:02X}".format(bytes_value)
    return tag_hex_value

def run_test():
    global latest_tag
    global tag_count
    global tag_data
    
    tag_bytes_list_for_device_1 = []
    tag_hex_value_list = set([])
    should_read_tag_from_device_1 = False

    try:
        serial_device_1 = serial.Serial('COM4', 57600, timeout=0.5)
    except serial.serialutil.SerialException as err:
        print('There was a problem while opening the ports for the reader')
        raise err

    try:
        serial_device_1.reset_input_buffer()
        # create a new workbook and sheet
        wb = Workbook()
        sheet = wb.active
        sheet.title = "RFID Data"
        sheet['A1'] = "Tag Value"
        sheet['B1'] = "Timestamp"
        row_num = 2
        written_tag_values = set()

        while True:
            tag_hex_value = ""
            read_bytes_from_device_1 = serial_device_1.read()
            int_value_from_device_1 = int.from_bytes(read_bytes_from_device_1, "big")

            if int_value_from_device_1 == 0x11:
                should_read_tag_from_device_1 = True

            if should_read_tag_from_device_1 is True:
                tag_bytes_list_for_device_1.append(int_value_from_device_1)

                if len(tag_bytes_list_for_device_1) == 18:
                    should_read_tag_from_device_1 = False
                    tag_hex_value = convert_tag_from_bytes_to_hex(tag_bytes_list_for_device_1)
                    tag_hex_value_list.add(tag_hex_value)
                    tag_bytes_list_for_device_1.clear()

                    # update latest_tag and tag_count
                    latest_tag = tag_hex_value
                    tag_count = len(tag_hex_value_list)

                    if tag_hex_value not in written_tag_values :
                        # write to excel sheet
                        sheet.cell(row=row_num, column=1).value = tag_hex_value
                        written_tag_values.add(tag_hex_value)
                        sheet.cell(row=row_num, column=2).value = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")                                          
                        row_num += 1
                        wb.save("RFID_Data.xlsx")

            print(f"RFID Tag count: {tag_count}, Latest tag: {latest_tag}")
                    
    except KeyboardInterrupt:
        print("Received keyboard interrupt in the RFID reader test program. Closing the ports and exiting the program")
        serial_device_1.flush()
        serial_device_1.reset_input_buffer()
        serial_device_1.close()

        

        sys.exit(0)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/stream')
def stream():
    def event_stream():
        while True:
            yield 'data: {{"latest_tag": "{0}", "tag_count": "{1}"}}\n\n'.format(latest_tag, tag_count)
    return app.response_class(event_stream(), mimetype='text/event-stream')

if __name__ == "__main__":
    run_test_thread = threading.Thread(target=run_test)
    run_test_thread.start()

    app.run(debug=True)


